# MCLeadProspecta Web - Memocash Solucoes - github.com/laossim

import os, re, time, uuid, threading, unicodedata, urllib.request as _url
import difflib
from datetime import date, datetime
from flask import Flask, render_template, request, jsonify, send_file

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "mcleadprospecta-dev")
JOBS = {}
JOBS_LOCK = threading.Lock()

BROWSER_ARGS = [
    "--no-sandbox","--disable-dev-shm-usage","--disable-gpu",
    "--disable-blink-features=AutomationControlled",
    "--disable-setuid-sandbox","--single-process","--lang=pt-BR",
]
CTX_ARGS = dict(
    locale="pt-BR", timezone_id="America/Sao_Paulo",
    viewport={"width":1280,"height":800},
    user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
)

def normalizar(t):
    return unicodedata.normalize("NFKD",t).encode("ASCII","ignore").decode().lower().strip()

# ── CORREÇÃO APRIMORADA ───────────────────────────────────────────────────────

CORRECOES = {
    # Nichos
    "hamburguer":"hamburguer","hamburger":"hamburguer","hamburgeria":"hamburgueria",
    "pizaria":"pizzaria","pizzeria":"pizzaria","pizzara":"pizzaria",
    "padaraia":"padaria","paderia":"padaria","pandaria":"padaria",
    "farmacia":"farmacia","farrmacia":"farmacia","farmasia":"farmacia",
    "acadmia":"academia","academai":"academia","adademia":"academia",
    "mecanica":"mecanica","mecanico":"mecanico","mecanika":"mecanica",
    "odontolgia":"odontologia","odontologya":"odontologia",
    "adovgado":"advogado","advogato":"advogado",
    "contabilidde":"contabilidade","contabiliade":"contabilidade",
    "restarante":"restaurante","restrurante":"restaurante","restaurente":"restaurante",
    "supermecado":"supermercado","supermecardo":"supermercado",
    "barberia":"barbearia","barbaria":"barbearia","barbiraria":"barbearia",
    "estetica":"estetica","esthetica":"estetica","estetica":"estetica",
    "petshop":"pet shop","pet shopt":"pet shop","petchop":"pet shop",
    "consulotrio":"consultorio","consultori":"consultorio",
    "clinika":"clinica","clincia":"clinica","klinica":"clinica",
    "sorvetaria":"sorveteria","sorvetaria":"sorveteria",
    "lanchinete":"lanchonete","lanchomete":"lanchonete",
    "mrecado":"mercado","ofcina":"oficina",
    "eletrica":"eletrica","hidraulica":"hidraulica",
    "imobilaria":"imobiliaria","hotal":"hotel","hotell":"hotel",
    "pousadda":"pousada","pusada":"pousada",
    "ballett":"ballet","balley":"ballet","bale":"ballet","balet":"ballet",
    "pilats":"pilates","pilate":"pilates",
    "ioga":"yoga","iogue":"yoga",
    "crossfite":"crossfit","crosfit":"crossfit",
    "salgadaria":"salgaderia","salgadaria":"salgaderia",
    "sorvete":"sorveteria","sorvetes":"sorveteria",
    "manicure":"manicure","manicuri":"manicure","manicuri":"manicure",
    "cabeleireiro":"cabeleireiro","cabeleiriro":"cabeleireiro",
    "fotografo":"fotografo","fotgrafo":"fotografo",
    "arquiteto":"arquiteto","arquiteto":"arquiteto",
    "engenheiro":"engenheiro","engenhero":"engenheiro",
    "psicologia":"psicologia","psicologya":"psicologia",
    "fisioterapia":"fisioterapia","fisioterapya":"fisioterapia",
    "nutricionista":"nutricionista","nutriscionista":"nutricionista",
    "personal":"personal trainer","personall":"personal trainer",
    "muay thai":"muay thai","muaythai":"muay thai","muay-thai":"muay thai",
    "jiu jitsu":"jiu-jitsu","jiujitsu":"jiu-jitsu","jiu-jitso":"jiu-jitsu",
    "karate":"karate","carate":"karate",
    "natacao":"natacao","natassao":"natacao",
    "ginastica":"ginastica","gimnastica":"ginastica",
    "dansa":"danca","danssa":"danca",
    "ballet classico":"ballet classico","balé classico":"ballet classico",
    # Cidades
    "sao paulo":"Sao Paulo","san paulo":"Sao Paulo","saõ paulo":"Sao Paulo",
    "rio de janerio":"Rio de Janeiro","rio de janiero":"Rio de Janeiro","riodejaneiro":"Rio de Janeiro",
    "belo orizonte":"Belo Horizonte","bello horizonte":"Belo Horizonte",
    "curitba":"Curitiba","curtiba":"Curitiba","curitibba":"Curitiba",
    "forteleza":"Fortaleza","fortalleza":"Fortaleza",
    "salvaldor":"Salvador","slavador":"Salvador",
    "manuas":"Manaus","manaos":"Manaus","manus":"Manaus",
    "recfie":"Recife","reciife":"Recife","reciffe":"Recife",
    "poro alegre":"Porto Alegre","porto alegri":"Porto Alegre",
    "goainia":"Goiania","goinia":"Goiania","goiania":"Goiania",
    "camppinas":"Campinas","campinnas":"Campinas","campinhas":"Campinas",
    "brasilia":"Brasilia","brazilia":"Brasilia","brazília":"Brasilia",
    "florianpolis":"Florianopolis","florianopollis":"Florianopolis",
    "vittoria":"Vitoria","vitoria":"Vitoria",
    "macapa":"Macapa","belem":"Belem","bellem":"Belem",
    "treresina":"Teresina","teresinaa":"Teresina",
    "natel":"Natal","natall":"Natal",
    "maceio":"Maceio","maceioo":"Maceio",
    "joao pessoa":"Joao Pessoa","joaopessoa":"Joao Pessoa",
    "poto velho":"Porto Velho","porto veljo":"Porto Velho",
    "palmaz":"Palmas","palmax":"Palmas",
    "campo grand":"Campo Grande","campo grandi":"Campo Grande",
    "cuiaba":"Cuiaba","cuiabba":"Cuiaba",
    "ribeirao preto":"Ribeirao Preto","riberao preto":"Ribeirao Preto",
    "uberlandia":"Uberlandia","uberlania":"Uberlandia",
    "sorocba":"Sorocaba","sorocabba":"Sorocaba",
    "sao jose dos campos":"Sao Jose dos Campos",
    "guaruhos":"Guarulhos","guaruhos":"Guarulhos",
    "joinvile":"Joinville","joinvilli":"Joinville",
    "londrinna":"Londrina","londrina":"Londrina",
    "maringa":"Maringa","maringga":"Maringa",
    "juiz de forra":"Juiz de Fora","juizde fora":"Juiz de Fora",
    "niteroi":"Niteroi","niterõi":"Niteroi",
    "nova iguacu":"Nova Iguacu","nova iguassu":"Nova Iguacu",
    "itabira":"Itabira","itabirra":"Itabira",
    "betim":"Betim","contagem":"Contagem","ipatinga":"Ipatinga",
}

def corrigir(texto):
    texto = texto.strip()
    chave = normalizar(texto)
    if chave in CORRECOES:
        return CORRECOES[chave]

    # Tenta por similaridade (difflib) — captura erros de digitacao nao mapeados
    matches = difflib.get_close_matches(chave, CORRECOES.keys(), n=1, cutoff=0.85)
    if matches:
        return CORRECOES[matches[0]]

    # Corrige palavra por palavra para frases compostas
    palavras = texto.strip().split()
    resultado = []
    for p in palavras:
        pn = normalizar(p)
        if pn in CORRECOES:
            resultado.append(CORRECOES[pn])
        else:
            m = difflib.get_close_matches(pn, CORRECOES.keys(), n=1, cutoff=0.82)
            resultado.append(CORRECOES[m[0]] if m else p)
    return " ".join(resultado)

# ── FILTRO DE RELEVÂNCIA ──────────────────────────────────────────────────────

STOP_WORDS = {"de","da","do","em","e","a","o","as","os","para","com","um","uma",
              "no","na","por","que","se","ao","aos","las","los","el","la"}

SINONIMOS = {
    "ballet":    ["ballet","bale","balet","danca","dance","studio","classico","royal"],
    "pilates":   ["pilates","studio","pilat"],
    "yoga":      ["yoga","yogi","studio","meditacao"],
    "natacao":   ["natacao","piscina","aqua","swim","nadar"],
    "danca":     ["danca","dance","ballet","samba","forro","ritmo","studio"],
    "ginastica": ["ginastica","gym","fitness","academia"],
    "academia":  ["academia","gym","fitness","crossfit","musculacao"],
    "pizzaria":  ["pizza","pizzaria","pizzeria","napolitana"],
    "barbearia": ["barber","barbearia","cabelo","corte","navalha"],
    "padaria":   ["padaria","bakery","paes","pao","confeitaria"],
    "farmacia":  ["farmacia","drogaria","remedio","medicamento"],
    "dentista":  ["dentista","odonto","dental","clinica","ortodontia"],
    "medico":    ["medico","clinica","saude","health","consultorio"],
    "advogado":  ["advogado","advocacia","juridico","direito","escritorio"],
    "contabilidade":["contabilidade","contabil","contador","fiscal"],
    "hotel":     ["hotel","pousada","hostel","inn","resort"],
    "restaurante":["restaurante","bistro","cozinha","food"],
    "petshop":   ["pet","animal","veterinaria","vet","patas"],
    "estetica":  ["estetica","beleza","beauty","spa","depilacao"],
    "salao":     ["salao","beleza","beauty","cabeleireiro","hair"],
    "sorveteria":["sorvete","sorveteria","gelato","ice cream","acai"],
    "lanchonete":["lanchonete","lanche","sanduiche","hamburguer","fast"],
    "mercado":   ["mercado","supermercado","mercearia","minimercado"],
    "mecanica":  ["mecanica","oficina","auto","carro","veiculo","motor"],
    "muay":      ["muay","thai","luta","arte marcial","esporte"],
    "jiu":       ["jiu","jitsu","luta","arte marcial","grappling"],
    "karate":    ["karate","luta","arte marcial","dojo"],
    "crossfit":  ["crossfit","functional","fitness","treino"],
    "personal":  ["personal","trainer","treino","fitness"],
}

def filtrar_relevantes(estabs, nicho):
    if not estabs:
        return estabs

    kws = [w for w in normalizar(nicho).split() if w not in STOP_WORDS and len(w) > 2]
    if not kws:
        return estabs

    # Expande com sinonimos
    expanded = set(kws)
    for kw in kws:
        if kw in SINONIMOS:
            expanded.update(SINONIMOS[kw])
        # Procura sinonimos por similaridade
        for sk, sv in SINONIMOS.items():
            if difflib.SequenceMatcher(None, kw, sk).ratio() > 0.8:
                expanded.update(sv)

    def relevante(nome):
        n = normalizar(nome)
        return any(kw in n for kw in expanded)

    filtrados = [e for e in estabs if relevante(e["nome"])]

    # Fallback: se filtrou mais de 60%, retorna tudo (nicho muito especifico)
    if len(filtrados) < len(estabs) * 0.4 and len(estabs) > 4:
        return estabs

    return filtrados if filtrados else estabs

def sanitizar(t):
    return re.sub(r"[^\w\s-]","",t).strip().replace(" ","_")

def pasta_output():
    p = os.path.join(os.path.dirname(__file__),"outputs")
    os.makedirs(p,exist_ok=True)
    return p

def gerar_whatsapp(tel):
    n = re.sub(r"\D","",tel)
    if not n: return ""
    if not n.startswith("55"): n="55"+n
    return "https://wa.me/"+n

PADRAO_TEL = re.compile(r"(?:\+?55\s?)?(?:\(?\d{2}\)?[\s\-]?)(?:9\s?)?\d{4}[\s\-]?\d{4}")

# ── EXTRAÇÃO ──────────────────────────────────────────────────────────────────

def extrair_detalhes(page, url):
    dados = {"endereco":"","telefone":"","whatsapp":""}
    try:
        page.goto(url, timeout=10000, wait_until="domcontentloaded")
        try:
            page.wait_for_selector(
                "button[aria-label*='Telefone'], button[aria-label*='telefone'], button[aria-label*='Phone']",
                timeout=4000
            )
        except Exception:
            pass

        botoes = page.locator("button[aria-label]")
        cnt = botoes.count()
        for i in range(min(cnt,30)):
            try:
                aria = botoes.nth(i).get_attribute("aria-label") or ""
                al = aria.lower()
                if not dados["endereco"] and any(k in al for k in ("endereco:","address:")):
                    dados["endereco"] = re.sub(r"^[^:]+:\s*","",aria).strip()
                if not dados["telefone"] and any(k in al for k in ("telefone:","phone:","numero")):
                    tel = re.sub(r"^[^:]+:\s*","",aria).strip()
                    if tel: dados["telefone"]=tel; dados["whatsapp"]=gerar_whatsapp(tel)
            except Exception:
                pass
            if dados["endereco"] and dados["telefone"]: return dados

        for cls in ("Io6YTe","rogA2c","AeaXub","LrzXr"):
            divs = page.locator("div."+cls)
            cnt = divs.count()
            for i in range(min(cnt,15)):
                try:
                    txt = divs.nth(i).inner_text().strip()
                    if not txt: continue
                    if not dados["endereco"] and len(txt)>10 and ","in txt and re.search(r"\d{4,}",txt):
                        if not re.search(r"[\+\(].*\d{4}",txt):
                            dados["endereco"]=txt.replace("\n",", ")
                    if not dados["telefone"] and re.match(r"^[\(\+\d][\d\s\(\)\-\.]{7,}$",txt):
                        dados["telefone"]=txt.strip(); dados["whatsapp"]=gerar_whatsapp(txt)
                except Exception:
                    pass
            if dados["endereco"] and dados["telefone"]: return dados

        try:
            corpo = page.inner_text("body")[:8000]
            if not dados["telefone"]:
                m = PADRAO_TEL.search(corpo)
                if m:
                    dados["telefone"]=m.group().strip()
                    dados["whatsapp"]=gerar_whatsapp(dados["telefone"])
        except Exception:
            pass
    except Exception:
        pass
    return dados

# ── PLANILHA ──────────────────────────────────────────────────────────────────

def salvar_xlsx(resultados, cidade, nicho, tag=""):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    def brd():
        s=Side(style="thin",color="a8d4ac")
        return Border(left=s,right=s,top=s,bottom=s)

    ws = wb.active; ws.title="Leads"
    ws.merge_cells("A1:H1")
    c=ws["A1"]; c.value="LEADS - "+nicho.upper()+" EM "+cidade.upper()
    c.font=Font(name="Arial",bold=True,size=14,color="0d4f18")
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.fill=PatternFill("solid",fgColor="d4edda")
    ws.row_dimensions[1].height=32

    ws.merge_cells("A2:H2")
    c2=ws["A2"]
    c2.value="Gerado em "+date.today().strftime("%d/%m/%Y")+" - "+str(len(resultados))+" leads - MCLeadProspecta"
    c2.font=Font(name="Arial",size=9,color="3a7d44",italic=True)
    c2.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=6

    for col,txt in enumerate(["Nome","Endereco","Telefone","WhatsApp","Link Maps","Status","Observacoes","Prioridade"],1):
        c=ws.cell(row=4,column=col,value=txt)
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="1e8a2e")
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=brd()
    ws.row_dimensions[4].height=24

    for idx,r in enumerate(resultados,1):
        lin=idx+4; cor="eaf5eb" if idx%2==0 else "FFFFFF"
        vals=[r["nome"],r["endereco"] or "-",r["telefone"] or "-",
              r["whatsapp"] or "-",r["url"],"Novo Lead","","Media"]
        for col,val in enumerate(vals,1):
            c=ws.cell(row=lin,column=col,value=val)
            c.font=Font(name="Arial",size=9)
            c.fill=PatternFill("solid",fgColor=cor)
            c.border=brd(); c.alignment=Alignment(vertical="center")
            if col==4 and val.startswith("http"):
                c.hyperlink=val; c.value="WhatsApp"
                c.font=Font(name="Arial",size=9,color="0d6b1e",underline="single")
            if col==5 and val.startswith("http"):
                c.hyperlink=val; c.value="Maps"
                c.font=Font(name="Arial",size=9,color="1e8a2e",underline="single")
            if col in(6,8): c.alignment=Alignment(horizontal="center",vertical="center")
            if col==6: c.font=Font(name="Arial",bold=True,size=9,color="1e8a2e")
        ws.row_dimensions[lin].height=17

    total=len(resultados)+4
    for sq,f1 in [
        ("F5:F"+str(total),'"Novo Lead,Tentativa de Contato,Contactado,Em Negociacao,Proposta Enviada,Fechado,Sem Interesse,Inativo"'),
        ("H5:H"+str(total),'"Alta,Media,Baixa,VIP"'),
    ]:
        dv=DataValidation(type="list",formula1=f1,allow_blank=True,showDropDown=False)
        dv.sqref=sq; ws.add_data_validation(dv)

    for col,w in zip("ABCDEFGH",[34,42,18,16,12,22,36,10]):
        ws.column_dimensions[col].width=w
    ws.freeze_panes="A5"

    sfx=("_"+tag) if tag else ""
    nome=sanitizar(cidade)+"_"+sanitizar(nicho)+"_"+date.today().strftime("%Y%m%d")+sfx+"_"+uuid.uuid4().hex[:6]+".xlsx"
    caminho=os.path.join(pasta_output(),nome)
    wb.save(caminho)
    return caminho,nome

# ── WORKER ────────────────────────────────────────────────────────────────────

MAX_MIN=25
SCROLL_JS="const f=document.querySelector(\"div[role='feed']\");if(f)f.scrollBy(0,2800);"

def worker(job_id, cidade, nicho):
    inicio=time.time()

    def log(msg):
        with JOBS_LOCK: JOBS[job_id]["log"].append(msg)
    def set_prog(v):
        with JOBS_LOCK: JOBS[job_id]["progress"]=v
    def set_stats(col,tel,atual,rate,eta):
        with JOBS_LOCK:
            JOBS[job_id].update({"coletados":col,"com_telefone":tel,"atual":atual,"rate":rate,"eta":eta})
    def deve_parar():
        with JOBS_LOCK: c=JOBS[job_id].get("cancelado",False)
        return c or (time.time()-inicio)>(MAX_MIN*60)

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log("Playwright nao instalado.")
        with JOBS_LOCK: JOBS[job_id]["status"]="erro"; return

    busca=nicho+" em "+cidade
    log("Iniciando: "+busca)

    try:
        with sync_playwright() as p:
            browser=p.chromium.launch(headless=True,args=BROWSER_ARGS)
            ctx=browser.new_context(**CTX_ARGS)

            pg=ctx.new_page(); pg.set_default_timeout(15000)
            pg.goto("https://www.google.com/maps/search/"+busca.replace(" ","+"),
                    timeout=50000,wait_until="domcontentloaded")

            try:
                pg.wait_for_selector("div[role='feed']",timeout=14000)
            except Exception:
                log("Google Maps nao respondeu.")
                with JOBS_LOCK: JOBS[job_id]["status"]="erro"
                browser.close(); return

            log("Carregando lista...")
            for _ in range(14):
                try: pg.evaluate(SCROLL_JS)
                except Exception: pass
                pg.wait_for_timeout(750)
                if pg.locator("span:has-text('fim da lista')").count()>0: break

            els=pg.locator("a[href*='/place/']")
            vistos=set(); estabs=[]
            for i in range(els.count()):
                try:
                    el=els.nth(i)
                    href=el.get_attribute("href") or ""
                    if "/place/" not in href: continue
                    nome=el.get_attribute("aria-label") or el.inner_text().strip()
                    if nome and len(nome)>2 and nome not in vistos and "google" not in nome.lower():
                        vistos.add(nome)
                        if href.startswith("/"): href="https://www.google.com"+href
                        estabs.append({"nome":nome,"url":href})
                except Exception: pass

            pg.close()

            # Aplica filtro de relevancia
            antes=len(estabs)
            estabs=filtrar_relevantes(estabs,nicho)
            filtrados=antes-len(estabs)
            total=len(estabs)
            msg=str(total)+" locais"
            if filtrados>0: msg+=" ("+str(filtrados)+" filtrados por relevancia)"
            log(msg+". Coletando contatos...")
            with JOBS_LOCK: JOBS[job_id]["total"]=total

            pg2=ctx.new_page(); pg2.set_default_timeout(12000)
            resultados=[]; t_ini=time.time()

            for idx,est in enumerate(estabs,1):
                if deve_parar():
                    log("Parando. Salvando "+str(len(resultados))+" leads..."); break

                det=extrair_detalhes(pg2,est["url"])
                r={"nome":est["nome"],"url":est["url"],
                   "telefone":det["telefone"],"whatsapp":det["whatsapp"],"endereco":det["endereco"]}
                resultados.append(r)

                com_tel=sum(1 for x in resultados if x["telefone"])
                elapsed=time.time()-t_ini
                rate=round((idx/elapsed)*60,1) if elapsed>0 else 0
                eta=int(((total-idx)/idx)*elapsed) if idx>0 else 0
                tag=" [tel]" if det["telefone"] else ""
                log("["+str(idx)+"/"+str(total)+"] "+est["nome"][:48]+tag)
                set_prog(int(idx/total*100))
                set_stats(idx,com_tel,est["nome"],rate,eta)

                if idx%20==0:
                    try:
                        cp,_=salvar_xlsx(resultados,cidade,nicho,"parcial"+str(idx))
                        with JOBS_LOCK: JOBS[job_id]["file_path_parcial"]=cp
                        log("Parcial: "+str(idx)+" leads, "+str(com_tel)+" com tel")
                    except Exception: pass

            pg2.close(); browser.close()

            if not resultados:
                log("Nenhum resultado.")
                with JOBS_LOCK: JOBS[job_id]["status"]="erro"; return

            caminho,nome_arq=salvar_xlsx(resultados,cidade,nicho)
            com_tel=sum(1 for r in resultados if r["telefone"])
            log("Concluido - "+str(len(resultados))+" leads - "+str(com_tel)+" com telefone")
            with JOBS_LOCK:
                JOBS[job_id].update({"status":"concluido","file_path":caminho,
                                     "file_name":nome_arq,"progress":100,
                                     "com_telefone":com_tel,"coletados":len(resultados)})
    except Exception as e:
        log("Erro: "+str(e))
        with JOBS_LOCK: JOBS[job_id]["status"]="erro"

# ── ROTAS ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index(): return render_template("index.html")

@app.route("/corrigir",methods=["POST"])
def rota_corrigir():
    d=request.get_json()
    return jsonify({"cidade":corrigir(d.get("cidade","").strip()),
                    "nicho":corrigir(d.get("nicho","").strip())})

@app.route("/iniciar",methods=["POST"])
def rota_iniciar():
    d=request.get_json()
    cidade=corrigir(d.get("cidade","").strip())
    nicho=corrigir(d.get("nicho","").strip())
    if not cidade or not nicho: return jsonify({"erro":"Preencha cidade e nicho"}),400
    jid=uuid.uuid4().hex
    with JOBS_LOCK:
        JOBS[jid]={"status":"rodando","log":[],"progress":0,
                   "file_path":None,"file_name":None,"file_path_parcial":None,
                   "total":0,"coletados":0,"com_telefone":0,
                   "atual":"","rate":0,"eta":0,"cancelado":False,
                   "cidade":cidade,"nicho":nicho,"criado_em":datetime.now().isoformat()}
    threading.Thread(target=worker,args=(jid,cidade,nicho),daemon=True).start()
    return jsonify({"job_id":jid,"cidade":cidade,"nicho":nicho})

@app.route("/cancelar/<jid>",methods=["POST"])
def rota_cancelar(jid):
    with JOBS_LOCK:
        if jid in JOBS: JOBS[jid]["cancelado"]=True
    return jsonify({"ok":True})

@app.route("/status/<jid>")
def rota_status(jid):
    with JOBS_LOCK: job=JOBS.get(jid)
    if not job: return jsonify({"erro":"nao encontrado"}),404
    return jsonify({"status":job["status"],"progress":job["progress"],
                    "log":job["log"][-40:],"total":job["total"],
                    "coletados":job.get("coletados",0),"com_telefone":job.get("com_telefone",0),
                    "atual":job.get("atual",""),"rate":job.get("rate",0),"eta":job.get("eta",0),
                    "file_name":job.get("file_name"),"tem_parcial":job.get("file_path_parcial") is not None})

@app.route("/download/<jid>")
def rota_download(jid):
    with JOBS_LOCK: job=JOBS.get(jid)
    if not job: return "nao encontrado",404
    caminho=job.get("file_path") or job.get("file_path_parcial")
    nome=job.get("file_name") or "leads_parcial.xlsx"
    if not caminho: return "arquivo nao disponivel",404
    return send_file(caminho,as_attachment=True,download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/reportar",methods=["POST"])
def rota_reportar():
    import json as _json
    d=request.get_json()
    msg=d.get("mensagem",""); log_txt=d.get("log","")
    webhook=os.environ.get("DISCORD_WEBHOOK","")
    if not webhook: return jsonify({"ok":False,"erro":"Webhook nao configurado"}),400
    payload={"content":"**Reporte MCLeadProspecta**","embeds":[{
        "title":"Erro reportado pelo usuario","color":0xe05252,
        "description":"**Mensagem:** "+msg[:300]+"\n```"+log_txt[-800:]+"```"}]}
    try:
        req=_url.Request(webhook,data=_json.dumps(payload).encode(),
                         headers={"Content-Type":"application/json"},method="POST")
        _url.urlopen(req,timeout=5)
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"ok":False,"erro":str(e)}),500

@app.route("/historico")
def rota_historico():
    with JOBS_LOCK:
        hist=[{"cidade":j["cidade"],"nicho":j["nicho"],"status":j["status"],
               "total":j["total"],"criado_em":j["criado_em"],"job_id":jid}
              for jid,j in JOBS.items()]
    hist.sort(key=lambda x:x["criado_em"],reverse=True)
    return jsonify(hist[:30])

if __name__=="__main__":
    app.run(host="0.0.0.0",port=int(os.environ.get("PORT",5000)),debug=False)
